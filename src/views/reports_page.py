"""
Página de Generación de Reportes
"""
import io
import os
import tempfile
from collections import Counter
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Sequence, Tuple

import matplotlib.pyplot as plt
import pandas as pd
import streamlit as st

# Importación opcional de FPDF para PDF
try:
    from fpdf import FPDF
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    FPDF = None


def _get_session_role() -> str:
    """Obtiene el rol activo dentro de la sesión."""
    user_data = st.session_state.get("user_data") or {}
    if st.session_state.get("user_id"):
        return user_data.get("role", "user")
    return "guest"


def _safe_parse_datetime(value: Any) -> Optional[datetime]:
    """Intenta convertir cualquier valor a datetime."""
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    try:
        return pd.to_datetime(value).to_pydatetime()
    except Exception:
        return None


def _build_week_windows(start_date, end_date, max_windows: int = 6) -> Tuple[List[Tuple[datetime, datetime]], List[str]]:
    """Crea ventanas semanales y etiquetas legibles."""
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())
    if start_dt > end_dt:
        start_dt, end_dt = end_dt, start_dt
    windows: List[Tuple[datetime, datetime]] = []
    cursor_start = start_dt
    while cursor_start <= end_dt:
        window_end = min(cursor_start + timedelta(days=6, hours=23, minutes=59, seconds=59), end_dt)
        windows.append((cursor_start, window_end))
        cursor_start = window_end + timedelta(seconds=1)
    if not windows:
        windows = [(start_dt, end_dt)]
    if len(windows) > max_windows:
        windows = windows[-max_windows:]
    labels = [f"{win_start.strftime('%d/%m')} - {win_end.strftime('%d/%m')}" for win_start, win_end in windows]
    return windows, labels


def _count_records_by_window(records: Sequence[Dict[str, Any]], date_field: str,
                             windows: Sequence[Tuple[datetime, datetime]]) -> List[int]:
    """Cuenta elementos por ventana temporal."""
    counts: List[int] = []
    for start, end in windows:
        total = 0
        for record in records:
            record_date = _safe_parse_datetime(record.get(date_field))
            if record_date and start <= record_date <= end:
                total += 1
        counts.append(total)
    return counts


def _sum_records_by_window(records: Sequence[Dict[str, Any]], date_field: str,
                           windows: Sequence[Tuple[datetime, datetime]], value_field: str) -> List[float]:
    """Suma un campo numérico por ventana temporal."""
    sums: List[float] = []
    for start, end in windows:
        total = 0.0
        for record in records:
            record_date = _safe_parse_datetime(record.get(date_field))
            if record_date and start <= record_date <= end:
                total += _parse_float(record.get(value_field))
        sums.append(total)
    return sums


def _extract_city_metadata(record: Dict[str, Any]) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """Devuelve (city_id, city_name, country) a partir de un registro con joins."""
    poi = record.get("points_of_interest") or record.get("poi") or {}
    city = poi.get("cities") or poi.get("city") or record.get("city") or {}
    city_id = poi.get("city_id") or city.get("id") or record.get("city_id")
    city_name = city.get("name") or record.get("city_name")
    country = city.get("country") or record.get("country")
    return city_id, city_name, country


def _matches_location(record: Dict[str, Any], city_id: Optional[str], country: Optional[str]) -> bool:
    """Verifica si un registro pertenece a la ubicación solicitada."""
    if not city_id and not country:
        return True
    record_city_id, _, record_country = _extract_city_metadata(record)
    if city_id and record_city_id != city_id:
        return False
    if country and record_country != country:
        return False
    return True


def _filter_by_date(records: Sequence[Dict[str, Any]], date_field: str,
                    start_dt: datetime, end_dt: datetime) -> List[Dict[str, Any]]:
    """Filtra registros por fecha."""
    filtered: List[Dict[str, Any]] = []
    for record in records:
        record_date = _safe_parse_datetime(record.get(date_field))
        if record_date and start_dt <= record_date <= end_dt:
            filtered.append(record)
    return filtered


def _apply_filters(records: Sequence[Dict[str, Any]], user_id: Optional[str],
                   city_id: Optional[str], country: Optional[str]) -> List[Dict[str, Any]]:
    """Aplica filtros de usuario y ubicación."""
    filtered: List[Dict[str, Any]] = []
    for record in records:
        record_user = record.get("user_id") or ((record.get("users") or {}).get("id"))
        if user_id and record_user != user_id:
            continue
        if not _matches_location(record, city_id, country):
            continue
        filtered.append(record)
    return filtered


def _parse_float(value: Any) -> float:
    """Convierte cualquier número en float seguro."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value))
    except (TypeError, ValueError):
        return 0.0


def _format_currency(value: float, currency: str = "€") -> str:
    """Formatea cantidades monetarias."""
    return f"{currency} {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _collect_admin_dataset(db, start_dt: datetime, end_dt: datetime,
                           user_id: Optional[str], city_id: Optional[str],
                           country: Optional[str]) -> Dict[str, List[Dict[str, Any]]]:
    """Obtiene visitas, reservas y estadísticas con los filtros indicados."""
    visits = db.get_visits_range(start_dt, end_dt) if hasattr(db, "get_visits_range") else []
    bookings = db.get_bookings_range(start_dt, end_dt) if hasattr(db, "get_bookings_range") else []
    stats = db.get_usage_stats_range(start_dt, end_dt) if hasattr(db, "get_usage_stats_range") else []
    filtered_visits = _apply_filters(visits, user_id, city_id, country)
    filtered_bookings = _apply_filters(bookings, user_id, city_id, country)
    filtered_stats = stats
    if user_id:
        filtered_stats = [stat for stat in stats if stat.get("user_id") == user_id]
    return {
        "visits": filtered_visits,
        "bookings": filtered_bookings,
        "stats": filtered_stats
    }

def show(db, n8n):
    """Muestra la página de reportes"""
    
    st.title("📄 Generador de Reportes")
    st.markdown("Genera reportes personalizados en formato PDF")
    
    if st.session_state.get("user_id") is None:
        st.info("Inicia sesión para generar reportes personalizados con tus datos.")
        return
    
    role = _get_session_role()
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        show_report_config(db, n8n, role)
    
    with col2:
        show_report_preview(role)


def show_report_config(db, n8n, role: str):
    """Muestra la configuración del reporte"""
    
    st.subheader("⚙️ Configuración del Reporte")
    is_admin = role == "admin"
    current_user_id = st.session_state.get("user_id")
    
    report_catalog = [
        "📊 Resumen General",
        "👤 Actividad de Usuario",
        "🏆 Análisis de Popularidad",
        "💰 Reporte Financiero",
        "📈 Tendencias y Estadísticas"
    ] if is_admin else ["👤 Actividad de Usuario"]
    
    report_type = st.selectbox(
        "Tipo de Reporte",
        report_catalog
    )
    
    if not is_admin:
        st.caption("Tus reportes estarán limitados a tu propia actividad dentro de la plataforma.")
    
    # Rango de fechas
    col1, col2 = st.columns(2)
    
    with col1:
        start_date = st.date_input(
            "Fecha de inicio",
            value=datetime.now().date() - timedelta(days=30)
        )
    
    with col2:
        end_date = st.date_input(
            "Fecha de fin",
            value=datetime.now().date()
        )
    
    selected_user_id = current_user_id
    selected_city_id = None
    selected_country = None
    
    if is_admin:
        st.markdown("### 🎛️ Filtros avanzados")
        
        user_records = db.get_all_users() if hasattr(db, "get_all_users") else []
        user_labels = ["Todos los usuarios"]
        user_lookup: Dict[str, str] = {}
        for user in user_records:
            label = f"{user.get('name', 'Sin nombre')} ({user.get('email', 'sin correo')})"
            user_labels.append(label)
            user_lookup[label] = user.get("id")
        selected_user_label = st.selectbox("Usuario", user_labels, index=0)
        if selected_user_label != "Todos los usuarios":
            selected_user_id = user_lookup.get(selected_user_label)
            st.caption(f"🔎 Filtrando datos para {selected_user_label}")
        else:
            selected_user_id = None
        
        cities_catalog = db.get_all_cities(include_inactive=True) if hasattr(db, "get_all_cities") else []
        countries = sorted({city.get("country") for city in cities_catalog if city.get("country")})
        country_options = ["Todos los países"] + countries
        selected_country_label = st.selectbox("País", country_options, index=0)
        if selected_country_label != "Todos los países":
            selected_country = selected_country_label
        
        city_entries = [("Todas las ciudades", None)]
        for city in cities_catalog:
            if selected_country and city.get("country") != selected_country:
                continue
            label = f"{city.get('name', 'Sin nombre')} ({city.get('country', 'N/D')})"
            city_entries.append((label, city.get("id")))
        city_labels = [entry[0] for entry in city_entries]
        selected_city_label = st.selectbox("Ciudad", city_labels, index=0)
        selected_city_id = next((entry[1] for entry in city_entries if entry[0] == selected_city_label), None)
    
    # Opciones adicionales
    st.markdown("### 📋 Opciones de Contenido")
    
    col1, col2 = st.columns(2)
    
    with col1:
        include_charts = st.checkbox("Incluir gráficos", value=True)
        include_tables = st.checkbox("Incluir tablas detalladas", value=True)
    
    with col2:
        include_summary = st.checkbox("Incluir resumen ejecutivo", value=True)
        include_recommendations = st.checkbox("Incluir recomendaciones", value=True)
    
    if "Usuario" in report_type:
        if not selected_user_id:
            st.warning("⚠️ Selecciona un usuario para continuar con este tipo de reporte.")
            return
        st.info(f"📧 Generando reporte para: {st.session_state.get('user_email') if selected_user_id == current_user_id else 'usuario seleccionado'}")
    
    # Botones de acción
    st.markdown("---")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("👁️ Vista Previa", type="secondary", use_container_width=True):
            generate_preview(
                db,
                report_type,
                start_date,
                end_date,
                include_charts,
                include_tables,
                include_summary,
                include_recommendations,
                selected_user_id,
                selected_city_id,
                selected_country,
                role
            )
    
    with col2:
        if st.button("📥 Generar PDF", type="primary", use_container_width=True):
            generate_pdf_report(
                db, n8n, report_type, start_date, end_date,
                include_charts, include_tables, include_summary, include_recommendations,
                selected_user_id,
                selected_city_id,
                selected_country,
                role
            )
    
    with col3:
        if st.button("📊 Generar Excel", type="primary", use_container_width=True):
            generate_excel_report(
                db, report_type, start_date, end_date,
                include_charts, include_tables, include_summary, include_recommendations,
                selected_user_id,
                selected_city_id,
                selected_country,
                role
            )


def show_report_preview(role):
    """Muestra una vista previa del reporte"""
    
    st.subheader("👁️ Vista Previa")
    
    if role == "admin":
        st.info("Configura el reporte y haz clic en 'Vista Previa' para ver el contenido. Como administrador, puedes filtrar por usuario, ciudad y país.")
    else:
        st.info("Configura el reporte y haz clic en 'Vista Previa' para ver el contenido. Tus reportes mostrarán únicamente tus datos personales.")
    
    # Placeholder para la vista previa
    with st.container():
        st.markdown("""
        ### 📊 Ejemplo de Reporte
        
        **Período:** 01/01/2024 - 31/01/2024
        
        #### Resumen Ejecutivo
        - Total de visitas: 1,247
        - Usuarios activos: 892
        - Reservas realizadas: 423
        - Ingresos totales: €15,247
        
        #### Métricas Principales
        - Tasa de conversión: 34%
        - Satisfacción promedio: 4.7/5.0
        - Tiempo promedio de visita: 45 min
        """)


def generate_preview(
    db,
    report_type,
    start_date,
    end_date,
    include_charts,
    include_tables,
    include_summary,
    include_recommendations,
    selected_user_id,
    selected_city_id,
    selected_country,
    role
):
    """Genera una vista previa del reporte"""
    
    with st.spinner("Generando vista previa..."):
        content = get_report_content(
            report_type,
            db,
            start_date,
            end_date,
            selected_user_id,
            selected_city_id,
            selected_country,
            role
        )
        st.success("Vista previa generada")
        
        with st.expander("📄 Vista Previa del Reporte", expanded=True):
            st.markdown(f"## {report_type}")
            st.markdown(f"**Período:** {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}")
            st.markdown(f"**Generado:** {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            
            st.markdown("---")
            render_report_preview(
                content,
                include_charts=include_charts,
                include_tables=include_tables,
                include_summary=include_summary,
                include_recommendations=include_recommendations
            )


def get_report_content(report_type, db, start_date, end_date,
                       selected_user_id=None, selected_city_id=None, selected_country=None, role="user"):
    """Construye la información necesaria para cada tipo de reporte"""
    
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())
    week_windows, week_labels = _build_week_windows(start_date, end_date)
    session_user_id = st.session_state.get("user_id")
    
    # Cache para dataset de admin
    dataset_cache = {}
    
    def admin_dataset():
        """Obtiene dataset filtrado para administradores"""
        if "value" not in dataset_cache:
            dataset_cache["value"] = _collect_admin_dataset(
                db, start_dt, end_dt, selected_user_id, selected_city_id, selected_country
            )
        return dataset_cache["value"]
    
    if "Resumen General" in report_type:
        if role != "admin":
            return {
                "summary_points": ["Este reporte está disponible solo para administradores."],
                "detail_items": [],
                "metrics": [],
                "chart": {"title": "", "data": pd.DataFrame(), "kind": "line"},
                "table": {"title": "", "data": pd.DataFrame()},
                "recommendations": []
            }
        
        dataset = admin_dataset()
        visits = dataset["visits"]
        bookings = dataset["bookings"]
        confirmed = [b for b in bookings if b.get("status") in ("confirmed", "completed")]
        pending = len([b for b in bookings if b.get("status") == "pending"])
        revenue = sum(_parse_float(b.get("total_price")) for b in confirmed)
        visits_weekly = _count_records_by_window(visits, "visit_date", week_windows)
        bookings_weekly = _count_records_by_window(bookings, "booking_date", week_windows)
        
        chart_df = pd.DataFrame({
            "Periodo": week_labels,
            "Visitas": visits_weekly,
            "Reservas": bookings_weekly
        })
        
        # Estadísticas por ciudad
        city_summary = {}
        for visit in visits:
            _, city_name, country = _extract_city_metadata(visit)
            key = city_name or "Sin ciudad"
            entry = city_summary.setdefault(key, {"País": country or "N/D", "Visitas": 0, "Reservas": 0})
            entry["Visitas"] += 1
        
        for booking in bookings:
            _, city_name, country = _extract_city_metadata(booking)
            key = city_name or "Sin ciudad"
            entry = city_summary.setdefault(key, {"País": country or "N/D", "Visitas": 0, "Reservas": 0})
            entry["Reservas"] += 1
        
        if city_summary:
            table_rows = [{
                "Ciudad": key,
                "País": data["País"],
                "Visitas": data["Visitas"],
                "Reservas": data["Reservas"]
            } for key, data in city_summary.items()]
            table_df = pd.DataFrame(table_rows).sort_values(by="Visitas", ascending=False).head(5)
        else:
            table_df = pd.DataFrame({
                "Ciudad": ["Sin datos"],
                "País": ["-"],
                "Visitas": [0],
                "Reservas": [0]
            })
        
        unique_users = len({
            visit.get("user_id") or ((visit.get("users") or {}).get("id"))
            for visit in visits
            if visit.get("user_id") or ((visit.get("users") or {}).get("id"))
        })
        
        avg_ticket = revenue / len(confirmed) if confirmed else 0
        
        summary_points = [
            f"Visitas registradas en el período: {len(visits)}.",
            f"Reservas totales: {len(bookings)} (confirmadas: {len(confirmed)}).",
            f"Ingresos estimados: {_format_currency(revenue)}."
        ]
        
        detail_items = [
            f"Usuarios únicos impactados: {unique_users or 0}.",
            f"Reservas pendientes: {pending}.",
            f"Ticket promedio: {_format_currency(avg_ticket)}."
        ]
        
        top_city = table_df.iloc[0] if not table_df.empty else None
        recommendations = [
            f"Refuerza campañas en {top_city['Ciudad']} para capitalizar sus {int(top_city['Visitas'])} visitas." if top_city is not None else "Promueve nuevos destinos.",
            f"Da seguimiento a las {pending} reservas pendientes para evitar cancelaciones.",
            "Comparte testimonios en los destinos con mejor valoración."
        ]
        
        return {
            "summary_points": summary_points,
            "detail_items": detail_items,
            "metrics": [
                {"label": "Visitas registradas", "value": str(len(visits)), "delta": f"{len(week_labels)} semanas"},
                {"label": "Reservas confirmadas", "value": str(len(confirmed)), "delta": f"{pending} pendientes"},
                {"label": "Ingresos estimados", "value": _format_currency(revenue), "delta": f"{len(confirmed)} operaciones"}
            ],
            "chart": {
                "title": "Evolución semanal de visitas y reservas",
                "data": chart_df,
                "kind": "line"
            },
            "table": {
                "title": "Desempeño por ciudad",
                "data": table_df
            },
            "recommendations": recommendations
        }
    
    if "Usuario" in report_type:
        target_user_id = selected_user_id or session_user_id
        if not target_user_id:
            return {
                "summary_points": ["No se encontró un usuario válido para generar el reporte."],
                "detail_items": [],
                "metrics": [],
                "chart": {
                    "title": "Actividad semanal del usuario",
                    "data": pd.DataFrame({"Periodo": week_labels, "Interacciones": [0] * len(week_labels)}),
                    "kind": "line"
                },
                "table": {"title": "Participación por categoría", "data": pd.DataFrame()},
                "recommendations": ["Inicia sesión o selecciona un usuario para continuar."]
            }
        
        # Obtener datos del usuario
        if selected_user_id and selected_user_id != session_user_id and hasattr(db, "get_user_by_id"):
            user_data = db.get_user_by_id(target_user_id) or {}
            user_email = user_data.get("email", "sin correo")
        else:
            user_data = getattr(st.session_state, "user_data", {}) or {}
            user_email = getattr(st.session_state, "user_email", "sin correo")
        
        visits = db.get_user_visits(target_user_id) if hasattr(db, "get_user_visits") else []
        bookings = db.get_user_bookings(target_user_id) if hasattr(db, "get_user_bookings") else []
        achievements = db.get_user_achievements(target_user_id) if hasattr(db, "get_user_achievements") else []
        
        # Filtrar por fecha y ubicación
        visits = _filter_by_date(visits, "visit_date", start_dt, end_dt)
        bookings = _filter_by_date(bookings, "booking_date", start_dt, end_dt)
        achievements = _filter_by_date(achievements, "earned_at", start_dt, end_dt)
        visits = [v for v in visits if _matches_location(v, selected_city_id, selected_country)]
        bookings = [b for b in bookings if _matches_location(b, selected_city_id, selected_country)]
        
        confirmed_bookings = [b for b in bookings if b.get("status") in ("confirmed", "completed")]
        points = user_data.get("total_points", 0)
        
        chart_df = pd.DataFrame({
            "Periodo": week_labels,
            "Interacciones": _count_records_by_window(visits, "visit_date", week_windows),
            "Reservas": _count_records_by_window(bookings, "booking_date", week_windows)
        })
        
        # Estadísticas por categoría
        category_stats = {}
        for visit in visits:
            poi = visit.get("points_of_interest") or {}
            category = poi.get("category") or "General"
            entry = category_stats.setdefault(category, {"visits": 0, "bookings": 0, "ratings": []})
            entry["visits"] += 1
            rating = visit.get("rating")
            if isinstance(rating, (int, float)):
                entry["ratings"].append(float(rating))
        
        for booking in bookings:
            poi = booking.get("points_of_interest") or {}
            category = poi.get("category") or "General"
            entry = category_stats.setdefault(category, {"visits": 0, "bookings": 0, "ratings": []})
            entry["bookings"] += 1
        
        if category_stats:
            table_rows = []
            for category, data in category_stats.items():
                ratings = data.get("ratings") or []
                avg_rating = sum(ratings) / len(ratings) if ratings else 0
                table_rows.append({
                    "Categoría": category,
                    "Visitas": data["visits"],
                    "Reservas": data["bookings"],
                    "Valoración": round(avg_rating, 2)
                })
            table_df = pd.DataFrame(table_rows).sort_values(by="Visitas", ascending=False)
        else:
            table_df = pd.DataFrame({
                "Categoría": ["Sin datos"],
                "Visitas": [0],
                "Reservas": [0],
                "Valoración": [0]
            })
        
        last_visit = visits[0] if visits else None
        
        summary_points = [
            f"Visitas registradas en el período: {len(visits)}.",
            f"Reservas confirmadas: {len(confirmed_bookings)}.",
            f"Puntos acumulados: {points}."
        ]
        
        if last_visit:
            poi = last_visit.get("points_of_interest") or {}
            poi_name = poi.get("name") or last_visit.get("poi_name") or "Actividad reciente"
            visit_date = _safe_parse_datetime(last_visit.get("visit_date"))
            summary_points.append(f"Última visita: {poi_name} ({visit_date.strftime('%d/%m/%Y') if visit_date else 'sin fecha'}).")
        
        detail_items = [
            f"Correo asociado: {user_email}",
            f"Logros obtenidos en el período: {len(achievements)}",
            f"Reservas pendientes: {len(bookings) - len(confirmed_bookings)}"
        ]
        
        recommendations = [
            "Planifica una nueva visita en la ciudad filtrada para mantener el ritmo." if visits else "Agenda tu primera visita en el período seleccionado.",
            "Comparte reseñas después de cada experiencia para mejorar tus recomendaciones.",
            "Aprovecha el saldo de puntos para desbloquear beneficios adicionales."
        ]
        
        return {
            "summary_points": summary_points,
            "detail_items": detail_items,
            "metrics": [
                {"label": "Lugares visitados", "value": str(len(visits)), "delta": f"{len(set(v.get('poi_id') for v in visits))} POIs únicos"},
                {"label": "Reservas confirmadas", "value": str(len(confirmed_bookings)), "delta": f"{len(bookings)} totales"},
                {"label": "Logros obtenidos", "value": str(len(achievements)), "delta": "Actualizados al rango seleccionado"}
            ],
            "chart": {
                "title": "Actividad semanal del usuario",
                "data": chart_df,
                "kind": "line"
            },
            "table": {
                "title": "Participación por categoría",
                "data": table_df
            },
            "recommendations": recommendations
        }
    
    if "Popularidad" in report_type:
        if role != "admin":
            return {
                "summary_points": ["Este reporte está disponible solo para administradores."],
                "detail_items": [],
                "metrics": [],
                "chart": {"title": "", "data": pd.DataFrame(), "kind": "bar"},
                "table": {"title": "", "data": pd.DataFrame()},
                "recommendations": []
            }
        
        dataset = admin_dataset()
        visits = dataset["visits"]
        bookings = dataset["bookings"]
        
        # Agrupar por POI
        poi_summary = {}
        for visit in visits:
            poi = visit.get("points_of_interest") or {}
            poi_id = visit.get("poi_id") or poi.get("id") or f"visit-{visit.get('id')}"
            entry = poi_summary.setdefault(poi_id, {
                "Lugar": poi.get("name") or "Sin nombre",
                "Ciudad": _extract_city_metadata(visit)[1] or "N/D",
                "Visitas": 0,
                "Reservas": 0,
                "ratings": []
            })
            entry["Visitas"] += 1
            rating = visit.get("rating")
            if isinstance(rating, (int, float)):
                entry["ratings"].append(float(rating))
        
        for booking in bookings:
            poi = booking.get("points_of_interest") or {}
            poi_id = booking.get("poi_id") or poi.get("id") or f"booking-{booking.get('id')}"
            entry = poi_summary.setdefault(poi_id, {
                "Lugar": poi.get("name") or "Sin nombre",
                "Ciudad": _extract_city_metadata(booking)[1] or "N/D",
                "Visitas": 0,
                "Reservas": 0,
                "ratings": []
            })
            entry["Reservas"] += 1
        
        rows = []
        for data in poi_summary.values():
            ratings = data.pop("ratings", [])
            avg_rating = sum(ratings) / len(ratings) if ratings else 0
            rows.append({
                "Lugar": data["Lugar"],
                "Ciudad": data["Ciudad"],
                "Visitas": data["Visitas"],
                "Reservas": data["Reservas"],
                "Valoración": round(avg_rating, 2)
            })
        
        if not rows:
            rows = [{"Lugar": "Sin datos", "Ciudad": "-", "Visitas": 0, "Reservas": 0, "Valoración": 0}]
        
        table_df = pd.DataFrame(rows).sort_values(by="Visitas", ascending=False).head(5)
        chart_df = pd.DataFrame({
            "Lugar": table_df["Lugar"],
            "Visitas": table_df["Visitas"]
        })
        
        top_place = table_df.iloc[0] if not table_df.empty else None
        
        summary_points = [
            f"{top_place['Lugar']} lidera con {int(top_place['Visitas'])} visitas en el período." if top_place is not None else "No hay datos suficientes.",
            f"Reservas generadas por el top 5: {int(table_df['Reservas'].sum())}.",
            f"Valoración media destacada: {table_df['Valoración'].mean():.2f} / 5."
        ]
        
        detail_items = [
            f"La ciudad más demandada: {top_place['Ciudad']}." if top_place else "Sin datos de ciudades.",
            f"Reservas promedio por atractivo: {table_df['Reservas'].mean():.1f}.",
            "Los atractivos con mejor valoración concentran más reseñas positivas."
        ]
        
        recommendations = [
            f"Refuerza la disponibilidad en {top_place['Lugar']} para capitalizar la demanda." if top_place else "Promueve nuevos atractivos.",
            "Cruza promociones entre los atractivos gastronómicos y culturales mejor valorados.",
            "Incorpora testimonios recientes en las fichas con mayor conversión."
        ]
        
        return {
            "summary_points": summary_points,
            "detail_items": detail_items,
            "metrics": [
                {"label": "Visitas promedio (Top 5)", "value": f"{table_df['Visitas'].mean():.1f}", "delta": f"{len(poi_summary)} lugares analizados"},
                {"label": "Reservas convertidas", "value": str(int(table_df['Reservas'].sum())), "delta": "Esperadas en el período"},
                {"label": "Valoración media", "value": f"{table_df['Valoración'].mean():.2f}", "delta": "Sobre 5 puntos"}
            ],
            "chart": {
                "title": "Visitas por atractivo destacado",
                "data": chart_df,
                "kind": "bar"
            },
            "table": {
                "title": "Top 5 atractivos",
                "data": table_df
            },
            "recommendations": recommendations
        }
    
    if "Financiero" in report_type:
        if role != "admin":
            return {
                "summary_points": ["Este reporte está disponible solo para administradores."],
                "detail_items": [],
                "metrics": [],
                "chart": {"title": "", "data": pd.DataFrame(), "kind": "line"},
                "table": {"title": "", "data": pd.DataFrame()},
                "recommendations": []
            }
        
        dataset = admin_dataset()
        bookings = dataset["bookings"]
        confirmed = [b for b in bookings if b.get("status") in ("confirmed", "completed")]
        revenue = sum(_parse_float(b.get("total_price")) for b in confirmed)
        avg_ticket = revenue / len(confirmed) if confirmed else 0
        refunded = len([b for b in bookings if b.get("status") in ("cancelled", "refunded")])
        pending = len([b for b in bookings if b.get("status") == "pending"])
        
        revenue_series = _sum_records_by_window(confirmed, "booking_date", week_windows, "total_price")
        bookings_series = _count_records_by_window(bookings, "booking_date", week_windows)
        
        chart_df = pd.DataFrame({
            "Periodo": week_labels,
            "Ingresos": revenue_series,
            "Reservas": bookings_series
        })
        
        # Estadísticas por ciudad
        revenue_by_city = {}
        for booking in confirmed:
            _, city_name, country = _extract_city_metadata(booking)
            key = city_name or "Sin ciudad"
            entry = revenue_by_city.setdefault(key, {"País": country or "N/D", "Ingresos": 0.0, "Reservas": 0})
            entry["Ingresos"] += _parse_float(booking.get("total_price"))
            entry["Reservas"] += 1
        
        if revenue_by_city:
            table_rows = [{
                "Ciudad": key,
                "País": data["País"],
                "Ingresos": _format_currency(data["Ingresos"]),
                "Reservas": data["Reservas"]
            } for key, data in revenue_by_city.items()]
            table_df = pd.DataFrame(table_rows).sort_values(by="Reservas", ascending=False)
        else:
            table_df = pd.DataFrame({
                "Ciudad": ["Sin datos"],
                "País": ["-"],
                "Ingresos": ["€ 0,00"],
                "Reservas": [0]
            })
        
        refund_rate = (refunded / len(bookings) * 100) if bookings else 0
        
        summary_points = [
            f"Ingresos totales en el período: {_format_currency(revenue)}.",
            f"Reservas confirmadas: {len(confirmed)} de {len(bookings)} totales.",
            f"Ticket promedio: {_format_currency(avg_ticket)}."
        ]
        
        detail_items = [
            f"Reservas pendientes: {pending}.",
            f"Cancelaciones/Reembolsos: {refunded} ({refund_rate:.1f}%).",
            f"Ciudades con ingresos: {len(revenue_by_city)}."
        ]
        
        recommendations = [
            "Optimiza las campañas de performance para sostener el crecimiento.",
            "Revisa las reservas pendientes para evitar cancelaciones.",
            "Introduce ofertas escalonadas para incrementar el ticket medio."
        ]
        
        return {
            "summary_points": summary_points,
            "detail_items": detail_items,
            "metrics": [
                {"label": "Ingresos totales", "value": _format_currency(revenue), "delta": f"{len(confirmed)} operaciones"},
                {"label": "Reservas pagadas", "value": str(len(confirmed)), "delta": f"{pending} pendientes"},
                {"label": "Ticket promedio", "value": _format_currency(avg_ticket), "delta": f"{len(confirmed)} reservas"}
            ],
            "chart": {
                "title": "Ingresos y reservas por semana",
                "data": chart_df,
                "kind": "line"
            },
            "table": {
                "title": "Desempeño por ciudad",
                "data": table_df
            },
            "recommendations": recommendations
        }
    
    # Tendencias y estadísticas
    if "Tendencias" in report_type or "Estadísticas" in report_type:
        if role != "admin":
            return {
                "summary_points": ["Este reporte está disponible solo para administradores."],
                "detail_items": [],
                "metrics": [],
                "chart": {"title": "", "data": pd.DataFrame(), "kind": "line"},
                "table": {"title": "", "data": pd.DataFrame()},
                "recommendations": []
            }
        
        dataset = admin_dataset()
        visits = dataset["visits"]
        bookings = dataset["bookings"]
        stats = dataset["stats"]
        
        # Usuarios únicos
        unique_users = set()
        for visit in visits:
            user_id = visit.get("user_id") or ((visit.get("users") or {}).get("id"))
            if user_id:
                unique_users.add(user_id)
        
        for booking in bookings:
            user_id = booking.get("user_id") or ((booking.get("users") or {}).get("id"))
            if user_id:
                unique_users.add(user_id)
        
        # Estadísticas por tipo de acción
        action_counts = Counter(stat.get("action_type") for stat in stats if stat.get("action_type"))
        
        # Usuarios nuevos vs recurrentes (simplificado)
        visits_by_week = _count_records_by_window(visits, "visit_date", week_windows)
        bookings_by_week = _count_records_by_window(bookings, "booking_date", week_windows)
        
        chart_df = pd.DataFrame({
            "Periodo": week_labels,
            "Visitas": visits_by_week,
            "Reservas": bookings_by_week
        })
        
        table_rows = [
            {"Indicador": "Visitas totales", "Valor": len(visits), "Tendencia": "↗︎" if len(visits) > 0 else "→"},
            {"Indicador": "Reservas totales", "Valor": len(bookings), "Tendencia": "↗︎" if len(bookings) > 0 else "→"},
            {"Indicador": "Usuarios únicos", "Valor": len(unique_users), "Tendencia": "↗︎" if len(unique_users) > 0 else "→"},
            {"Indicador": "Acciones registradas", "Valor": len(stats), "Tendencia": "↗︎" if len(stats) > 0 else "→"}
        ]
        
        table_df = pd.DataFrame(table_rows)
        
        top_actions = action_counts.most_common(3)
        top_action_names = [action[0] for action in top_actions]
        
        summary_points = [
            f"Usuarios únicos activos en el período: {len(unique_users)}.",
            f"Total de interacciones registradas: {len(stats)}.",
            f"Acciones más frecuentes: {', '.join(top_action_names[:2])}."
        ]
        
        detail_items = [
            f"Visitas registradas: {len(visits)}.",
            f"Reservas procesadas: {len(bookings)}.",
            f"Tipos de acciones diferentes: {len(action_counts)}."
        ]
        
        recommendations = [
            "Prioriza el onboarding interactivo para mantener la tasa de activación.",
            "Refuerza las notificaciones segmentadas con base en intereses recientes.",
            "Aumenta el catálogo de experiencias para retener usuarios."
        ]
        
        return {
            "summary_points": summary_points,
            "detail_items": detail_items,
            "metrics": [
                {"label": "Usuarios únicos", "value": str(len(unique_users)), "delta": f"{len(visits)} visitas"},
                {"label": "Interacciones", "value": str(len(stats)), "delta": f"{len(action_counts)} tipos"},
                {"label": "Reservas", "value": str(len(bookings)), "delta": f"{len([b for b in bookings if b.get('status') in ('confirmed', 'completed')])} confirmadas" if bookings else "0"}
            ],
            "chart": {
                "title": "Evolución de visitas y reservas",
                "data": chart_df,
                "kind": "line"
            },
            "table": {
                "title": "Indicadores clave del período",
                "data": table_df
            },
            "recommendations": recommendations
        }
    
    # Fallback por defecto
    return {
        "summary_points": ["Tipo de reporte no reconocido."],
        "detail_items": [],
        "metrics": [],
        "chart": {"title": "", "data": pd.DataFrame(), "kind": "line"},
        "table": {"title": "", "data": pd.DataFrame()},
        "recommendations": []
    }


def render_report_preview(content, include_charts, include_tables, include_summary, include_recommendations):
    """Renderiza la vista previa utilizando la configuración seleccionada"""
    
    metrics = content.get("metrics", [])
    if metrics:
        cols = st.columns(len(metrics))
        for col, metric in zip(cols, metrics):
            delta = metric.get("delta")
            description = metric.get("description")
            if delta:
                col.metric(metric["label"], metric["value"], delta)
            elif description:
                col.metric(metric["label"], metric["value"], description)
            else:
                col.metric(metric["label"], metric["value"])
    
    detail_items = content.get("detail_items", [])
    if detail_items:
        st.markdown("### 📋 Detalles Clave")
        for item in detail_items:
            st.write(f"- {item}")
    
    if include_summary:
        summary_points = content.get("summary_points", [])
        if summary_points:
            st.markdown("### 📝 Resumen Ejecutivo")
            for point in summary_points:
                st.write(f"- {point}")
    
    if include_charts:
        chart_info = content.get("chart")
        if chart_info and isinstance(chart_info, dict):
            chart_df = chart_info.get("data")
            if isinstance(chart_df, pd.DataFrame) and not chart_df.empty:
                st.markdown(f"### {chart_info.get('title', 'Visualización')}")
                df_to_plot = chart_df.set_index(chart_df.columns[0])
                kind = chart_info.get("kind", "line")
                if kind == "bar":
                    st.bar_chart(df_to_plot)
                elif kind == "area":
                    st.area_chart(df_to_plot)
                else:
                    st.line_chart(df_to_plot)
    
    if include_tables:
        table_info = content.get("table")
        if table_info and isinstance(table_info, dict):
            table_df = table_info.get("data")
            if isinstance(table_df, pd.DataFrame) and not table_df.empty:
                st.markdown(f"### {table_info.get('title', 'Tabla de detalle')}")
                st.dataframe(table_df, use_container_width=True, hide_index=True)
    
    if include_recommendations:
        recommendations = content.get("recommendations", [])
        if recommendations:
            st.markdown("### 💡 Recomendaciones")
            for rec in recommendations:
                st.write(f"- {rec}")


def generate_pdf_report(db, n8n, report_type, start_date, end_date, 
                       include_charts, include_tables, include_summary, include_recommendations,
                       selected_user_id, selected_city_id, selected_country, role):
    """Genera el reporte en formato PDF"""
    
    if not PDF_AVAILABLE:
        st.error("❌ El módulo fpdf no está instalado. Por favor instala con: pip install fpdf")
        st.info("💡 Alternativa: Usa la opción 'Generar Excel' para exportar el reporte.")
        return
    
    with st.spinner("📄 Generando reporte PDF..."):
        content = get_report_content(
            report_type,
            db,
            start_date,
            end_date,
            selected_user_id,
            selected_city_id,
            selected_country,
            role
        )
        
        pdf = PDFReport()
        pdf.set_auto_page_break(auto=True, margin=20)
        pdf.add_page()
        
        pdf.set_text_color(33, 33, 33)
        pdf.set_font("Arial", 'B', 20)
        pdf.cell(0, 12, safe_pdf_text("Guía Turística Virtual"), 0, 1, 'C')
        pdf.ln(2)
        
        pdf.set_font("Arial", 'B', 16)
        pdf.set_text_color(55, 71, 79)
        pdf.cell(0, 10, safe_pdf_text(report_type), 0, 1, 'C')
        pdf.ln(4)
        
        pdf.set_text_color(97, 97, 97)
        pdf.set_font("Arial", size=10)
        pdf.cell(0, 6, safe_pdf_text(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"), 0, 1)
        pdf.cell(0, 6, safe_pdf_text(f"Período: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"), 0, 1)
        pdf.ln(8)
        
        metrics = content.get("metrics", [])
        if metrics:
            add_pdf_section_header(pdf, "Indicadores principales", (25, 118, 210))
            add_pdf_metrics(pdf, metrics)
        
        detail_items = content.get("detail_items", [])
        if detail_items:
            add_pdf_section_header(pdf, "Detalles clave", (84, 110, 122))
            add_pdf_bullet_list(pdf, detail_items)
        
        if include_summary:
            summary_points = content.get("summary_points", [])
            if summary_points:
                add_pdf_section_header(pdf, "Resumen ejecutivo", (30, 136, 229))
                add_pdf_bullet_list(pdf, summary_points)
        
        chart_info = content.get("chart")
        if include_charts and chart_info:
            add_pdf_chart(
                pdf,
                chart_info.get("title", "Visualización"),
                chart_info.get("data"),
                chart_info.get("kind", "line")
            )
        
        table_info = content.get("table")
        if include_tables and table_info:
            add_pdf_table(
                pdf,
                table_info.get("title", "Tabla de detalle"),
                table_info.get("data")
            )
        
        if include_recommendations:
            recommendations = content.get("recommendations", [])
            if recommendations:
                add_pdf_section_header(pdf, "Recomendaciones", (56, 142, 60))
                add_pdf_bullet_list(pdf, recommendations)
        
        pdf_output = io.BytesIO()
        pdf_bytes = pdf.output(dest='S').encode('latin1')
        pdf_output.write(pdf_bytes)
        pdf_output.seek(0)
        
        # Botón de descarga
        st.success("✅ Reporte generado exitosamente!")
        
        st.download_button(
            label="📥 Descargar Reporte PDF",
            data=pdf_output,
            file_name=f"reporte_turismo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            mime="application/pdf",
            use_container_width=True
        )
        
        st.balloons()


def generate_excel_report(db, report_type, start_date, end_date,
                         include_charts, include_tables, include_summary, include_recommendations,
                         selected_user_id, selected_city_id, selected_country, role):
    """Genera el reporte en formato Excel"""
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        EXCEL_AVAILABLE = True
    except ImportError:
        st.error("❌ El módulo openpyxl no está instalado. Por favor instala con: pip install openpyxl")
        return
    
    with st.spinner("📊 Generando reporte Excel..."):
        content = get_report_content(
            report_type,
            db,
            start_date,
            end_date,
            selected_user_id,
            selected_city_id,
            selected_country,
            role
        )
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16)
        subtitle_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        
        # Título
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = "Guía Turística Virtual"
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Tipo de reporte
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = report_type
        cell.font = subtitle_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Información del reporte
        row += 1
        ws[f'A{row}'] = "Generado:"
        ws[f'B{row}'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        row += 1
        ws[f'A{row}'] = "Período:"
        ws[f'B{row}'] = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        row += 2
        
        # Métricas
        metrics = content.get("metrics", [])
        if metrics:
            ws[f'A{row}'] = "Indicadores Principales"
            ws[f'A{row}'].font = subtitle_font
            row += 1
            
            # Encabezados
            ws[f'A{row}'] = "Métrica"
            ws[f'B{row}'] = "Valor"
            ws[f'C{row}'] = "Variación"
            for col in ['A', 'B', 'C']:
                cell = ws[f'{col}{row}']
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            row += 1
            
            # Datos
            for metric in metrics:
                ws[f'A{row}'] = metric.get("label", "")
                ws[f'B{row}'] = metric.get("value", "")
                ws[f'C{row}'] = metric.get("delta", "")
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].border = border
                row += 1
            row += 1
        
        # Resumen ejecutivo
        if include_summary:
            summary_points = content.get("summary_points", [])
            if summary_points:
                ws[f'A{row}'] = "Resumen Ejecutivo"
                ws[f'A{row}'].font = subtitle_font
                row += 1
                for point in summary_points:
                    ws[f'A{row}'] = f"• {point}"
                    ws.merge_cells(f'A{row}:E{row}')
                    row += 1
                row += 1
        
        # Detalles clave
        detail_items = content.get("detail_items", [])
        if detail_items:
            ws[f'A{row}'] = "Detalles Clave"
            ws[f'A{row}'].font = subtitle_font
            row += 1
            for item in detail_items:
                ws[f'A{row}'] = f"• {item}"
                ws.merge_cells(f'A{row}:E{row}')
                row += 1
            row += 1
        
        # Tabla de datos
        if include_tables:
            table_info = content.get("table")
            if table_info and isinstance(table_info, dict):
                table_df = table_info.get("data")
                if isinstance(table_df, pd.DataFrame) and not table_df.empty:
                    ws[f'A{row}'] = table_info.get("title", "Tabla de detalle")
                    ws[f'A{row}'].font = subtitle_font
                    row += 1
                    
                    # Escribir DataFrame
                    for col_idx, col_name in enumerate(table_df.columns, start=1):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.value = col_name
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                    
                    row += 1
                    
                    for _, df_row in table_df.iterrows():
                        for col_idx, value in enumerate(df_row, start=1):
                            cell = ws.cell(row=row, column=col_idx)
                            cell.value = value
                            cell.border = border
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        row += 1
                    row += 1
        
        # Recomendaciones
        if include_recommendations:
            recommendations = content.get("recommendations", [])
            if recommendations:
                ws[f'A{row}'] = "Recomendaciones"
                ws[f'A{row}'].font = subtitle_font
                row += 1
                for rec in recommendations:
                    ws[f'A{row}'] = f"• {rec}"
                    ws.merge_cells(f'A{row}:E{row}')
                    row += 1
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Guardar en BytesIO
        excel_output = io.BytesIO()
        wb.save(excel_output)
        excel_output.seek(0)
        
        # Botón de descarga
        st.success("✅ Reporte Excel generado exitosamente!")
        
        st.download_button(
            label="📊 Descargar Reporte Excel",
            data=excel_output,
            file_name=f"reporte_turismo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        st.balloons()


def safe_pdf_text(value):
    """Asegura compatibilidad de caracteres con el PDF."""
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    return value.encode('latin-1', 'replace').decode('latin-1')


def add_pdf_section_header(pdf, title, fill_color):
    """Agrega un encabezado de sección con fondo de color."""
    pdf.set_fill_color(*fill_color)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(0, 9, safe_pdf_text(title), 0, 1, 'L', True)
    pdf.ln(2)
    pdf.set_text_color(33, 33, 33)


def add_pdf_metrics(pdf, metrics):
    """Dibuja una grilla de métricas destacadas."""
    if not metrics:
        return
    
    usable_width = pdf.w - pdf.l_margin - pdf.r_margin
    max_cols = 3
    rows = [metrics[i:i + max_cols] for i in range(0, len(metrics), max_cols)]
    
    for row in rows:
        col_width = usable_width / len(row)
        
        pdf.set_fill_color(240, 244, 252)
        pdf.set_font("Arial", 'B', 10)
        for metric in row:
            pdf.cell(col_width, 8, safe_pdf_text(metric.get("label", "")), border=1, align='L', fill=True)
        pdf.ln()
        
        pdf.set_font("Arial", 'B', 14)
        pdf.set_text_color(33, 33, 33)
        for metric in row:
            pdf.cell(col_width, 10, safe_pdf_text(metric.get("value", "")), border=1, align='L')
        pdf.ln()
        
        pdf.set_font("Arial", '', 9)
        for metric in row:
            delta = metric.get("delta") or metric.get("description") or ""
            if delta:
                if isinstance(delta, str) and delta.strip().startswith("-"):
                    pdf.set_text_color(198, 40, 40)
                else:
                    pdf.set_text_color(46, 125, 50)
            else:
                pdf.set_text_color(120, 120, 120)
            pdf.cell(col_width, 8, safe_pdf_text(delta), border=1, align='L')
        pdf.ln(10)
        pdf.set_text_color(33, 33, 33)


def add_pdf_bullet_list(pdf, items):
    """Agrega una lista con viñetas."""
    if not items:
        return
    
    pdf.set_font("Arial", size=11)
    pdf.set_text_color(55, 71, 79)
    for item in items:
        pdf.multi_cell(0, 6, safe_pdf_text(f"• {item}"))
    pdf.ln(4)
    pdf.set_text_color(33, 33, 33)


def add_pdf_table(pdf, title, dataframe):
    """Renderiza una tabla en el PDF."""
    if not isinstance(dataframe, pd.DataFrame) or dataframe.empty:
        return
    
    add_pdf_section_header(pdf, title, (39, 76, 119))
    
    pdf.set_font("Arial", 'B', 10)
    pdf.set_text_color(33, 33, 33)
    
    usable_width = pdf.w - pdf.l_margin - pdf.r_margin
    column_count = len(dataframe.columns)
    col_width = usable_width / column_count
    
    pdf.set_fill_color(224, 235, 255)
    for column in dataframe.columns:
        pdf.cell(col_width, 8, safe_pdf_text(column), border=1, align='C', fill=True)
    pdf.ln()
    
    pdf.set_font("Arial", size=10)
    pdf.set_fill_color(255, 255, 255)
    
    for _, row in dataframe.iterrows():
        for value in row:
            pdf.cell(col_width, 8, safe_pdf_text(value), border=1, align='C')
        pdf.ln()
    
    pdf.ln(6)


def add_pdf_chart(pdf, title, dataframe, kind):
    """Agrega una visualización como imagen al PDF."""
    if not isinstance(dataframe, pd.DataFrame) or dataframe.empty:
        return
    
    add_pdf_section_header(pdf, title, (38, 166, 154))
    
    fig, ax = plt.subplots(figsize=(6, 3))
    x_values = dataframe.iloc[:, 0]
    
    if kind == "bar":
        values = dataframe.iloc[:, 1]
        ax.bar(x_values, values, color="#1e88e5")
        ax.set_ylim(0, max(values) * 1.2 if len(values) else 1)
    else:
        for column in dataframe.columns[1:]:
            ax.plot(x_values, dataframe[column], marker='o', linewidth=2, label=column)
        if len(dataframe.columns) > 2:
            ax.legend(loc='upper left')
    
    ax.set_xlabel("")
    ax.set_ylabel("")
    ax.grid(True, linestyle='--', alpha=0.4)
    plt.xticks(rotation=35, ha='right')
    fig.tight_layout()
    
    temp_image = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
    temp_path = temp_image.name
    temp_image.close()
    
    fig.savefig(temp_path, dpi=220, bbox_inches='tight')
    plt.close(fig)
    
    pdf.image(temp_path, w=pdf.w - pdf.l_margin - pdf.r_margin)
    os.unlink(temp_path)
    pdf.ln(6)

class PDFReport(FPDF):
    """Clase personalizada para generar reportes PDF"""
    
    def header(self):
        """Encabezado del PDF"""
        self.set_fill_color(25, 118, 210)
        self.rect(0, 0, self.w, 18, 'F')
        self.set_y(6)
        self.set_text_color(255, 255, 255)
        self.set_font('Arial', 'B', 12)
        self.cell(0, 8, safe_pdf_text('Guía Turística Virtual - Reporte'), 0, 1, 'C')
        self.ln(4)
        self.set_text_color(33, 33, 33)
    
    def footer(self):
        """Pie de página del PDF"""
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(120, 120, 120)
        self.cell(0, 10, safe_pdf_text(f'Página {self.page_no()}'), 0, 0, 'C')
